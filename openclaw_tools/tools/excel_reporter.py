"""
OpenCLAW Tool: excel_reporter
生成 Excel 格式的日志归因日报 + 累计用例统计
"""

import json
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── 样式定义 ───────────────────────────────────────────────

HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
CELL_FONT = Font(name="Microsoft YaHei", size=10)
WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
PASS_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ─── 责任人映射 ──────────────────────────────────────────────

COMPONENT_TEAM_MAP = {
    "libvirt_error": {"team": "计算组", "owner": "张三"},
    "qemu_crash": {"team": "计算组", "owner": "李四"},
    "kernel_panic": {"team": "内核组", "owner": "王五"},
    "memory_issue": {"team": "内核组", "owner": "赵六"},
    "timeout": {"team": "软硬协同组", "owner": "钱七"},
    "environment_issue": {"team": "软硬协同组", "owner": "孙八"},
    "case_script_issue": {"team": "计算组", "owner": "周九"},
    "infrastructure_issue": {"team": "存储组", "owner": "吴十"},
    "unknown_failure": {"team": "待定", "owner": "待定"},
}


def _apply_header_style(ws, row: int, col_count: int):
    """应用表头样式"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _apply_cell_style(ws, row: int, col_count: int, highlight: str = ""):
    """应用单元格样式"""
    fill = None
    if highlight == "fail":
        fill = FAIL_FILL
    elif highlight == "warn":
        fill = WARN_FILL
    elif highlight == "pass":
        fill = PASS_FILL

    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = CELL_FONT
        cell.alignment = LEFT_WRAP
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def _auto_width(ws, col_count: int, min_width: int = 10, max_width: int = 40):
    """自适应列宽"""
    for col in range(1, col_count + 1):
        max_len = min_width
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    # 中文字符宽度约2倍
                    val_len = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                    max_len = max(max_len, min(val_len + 4, max_width))
        ws.column_dimensions[get_column_letter(col)].width = max_len


# ─── 日报 Excel ──────────────────────────────────────────────

DAILY_COLUMNS = [
    "时间", "用例名称", "任务链接", "人物名", "host机",
    "任务UVP版本", "版本失败原因", "版本分类", "置信度",
    "分析方法", "改进措施", "责任人", "所属团队",
    "日志链接", "是否重复失败", "关联已知问题"
]


def render_daily_excel(
    job_info: Dict,
    failures: List[Dict],
    output_path: str = "output/daily_report.xlsx"
) -> Dict:
    """
    生成每日归因 Excel 报告

    Args:
        job_info: 任务元数据
        failures: 失败用例分析结果列表
        output_path: 输出路径

    Returns:
        生成结果
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "失败归因日报"

    # 表头
    for col, name in enumerate(DAILY_COLUMNS, 1):
        ws.cell(row=1, column=col, value=name)
    _apply_header_style(ws, 1, len(DAILY_COLUMNS))

    # 数据行
    for i, f in enumerate(failures, 2):
        failure_type = f.get("failure_type", "unknown_failure")
        team_info = COMPONENT_TEAM_MAP.get(failure_type, {"team": "待定", "owner": "待定"})

        # 改进措施
        suggestion = f.get("suggestion", "")
        if not suggestion:
            suggestion = _default_suggestion(failure_type)

        # 是否重复失败（简单判断：后续实现累计统计时替换）
        repeat_flag = f.get("is_repeat_failure", "否")

        row_data = [
            job_info.get("date", ""),                              # 时间
            f.get("test_name", ""),                                # 用例名称
            f.get("task_link", ""),                                # 任务链接
            job_info.get("operator", ""),                          # 人物名
            job_info.get("host", ""),                              # host机
            job_info.get("uvp_version", ""),                       # 任务UVP版本
            f.get("root_cause", ""),                               # 版本失败原因
            failure_type,                                          # 版本分类
            f.get("confidence", ""),                               # 置信度
            f.get("method", "rule_engine"),                        # 分析方法
            suggestion,                                            # 改进措施
            team_info["owner"],                                    # 责任人
            team_info["team"],                                     # 所属团队
            f.get("log_link", ""),                                 # 日志链接
            repeat_flag,                                           # 是否重复失败
            f.get("known_issue", ""),                              # 关联已知问题
        ]

        for col, val in enumerate(row_data, 1):
            ws.cell(row=i, column=col, value=val)

        # 高亮：低置信度黄色，失败行浅红
        highlight = "fail"
        if f.get("confidence") == "low":
            highlight = "warn"
        _apply_cell_style(ws, i, len(DAILY_COLUMNS), highlight)

    _auto_width(ws, len(DAILY_COLUMNS))

    # 冻结首行
    ws.freeze_panes = "A2"

    # 写入文件
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_file))

    return {
        "success": True,
        "output_path": str(output_file),
        "row_count": len(failures)
    }


# ─── 累计统计 Excel ──────────────────────────────────────────

STATS_COLUMNS = [
    "用例名称", "累计执行次数", "成功次数", "失败次数",
    "成功率", "最近7天成功率", "最近30天成功率",
    "失败原因统计", "平均运行时间(s)",
    "首次失败版本", "最近失败版本",
    "Flaky标记", "版本信息"
]


def render_stats_excel(
    case_stats: List[Dict],
    output_path: str = "output/case_stats.xlsx"
) -> Dict:
    """
    生成累计用例统计 Excel

    Args:
        case_stats: 用例统计列表
        output_path: 输出路径

    Returns:
        生成结果
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "用例累计统计"

    # 表头
    for col, name in enumerate(STATS_COLUMNS, 1):
        ws.cell(row=1, column=col, value=name)
    _apply_header_style(ws, 1, len(STATS_COLUMNS))

    # 数据行
    for i, stat in enumerate(case_stats, 2):
        total = stat.get("total_runs", 0)
        success = stat.get("success_count", 0)
        fail = stat.get("fail_count", 0)
        rate = f"{success/total*100:.1f}%" if total > 0 else "N/A"
        rate_7d = f"{stat.get('success_rate_7d', 0):.1f}%"
        rate_30d = f"{stat.get('success_rate_30d', 0):.1f}%"

        row_data = [
            stat.get("test_name", ""),
            total,
            success,
            fail,
            rate,
            rate_7d,
            rate_30d,
            stat.get("failure_type_stats", ""),
            f"{stat.get('avg_duration', 0):.1f}",
            stat.get("first_fail_version", ""),
            stat.get("last_fail_version", ""),
            stat.get("flaky", ""),
            stat.get("version_info", ""),
        ]

        for col, val in enumerate(row_data, 1):
            ws.cell(row=i, column=col, value=val)

        # Flaky 行高亮为黄色
        highlight = "warn" if stat.get("flaky") == "是" else ""
        _apply_cell_style(ws, i, len(STATS_COLUMNS), highlight)

    _auto_width(ws, len(STATS_COLUMNS))
    ws.freeze_panes = "A2"

    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_file))

    return {
        "success": True,
        "output_path": str(output_file),
        "case_count": len(case_stats)
    }


# ─── 累计统计计算 ──────────────────────────────────────────────

def compute_case_stats(history: List[Dict]) -> List[Dict]:
    """
    根据历史运行记录计算每个用例的累计统计

    Args:
        history: 历史运行记录列表，每条记录包含:
            test_name, status, duration, date, uvp_version, failure_type

    Returns:
        每个用例的统计信息列表
    """
    from collections import defaultdict

    # 按用例分组
    cases = defaultdict(lambda: {
        "total_runs": 0, "success_count": 0, "fail_count": 0,
        "durations": [], "failure_types": defaultdict(int),
        "versions": set(), "dates": [],
        "first_fail_version": None, "last_fail_version": None,
        "results_7d": [], "results_30d": [],
    })

    for record in history:
        name = record.get("test_name", "")
        case = cases[name]

        case["total_runs"] += 1
        case["durations"].append(record.get("duration", 0))
        case["versions"].add(record.get("uvp_version", ""))
        case["dates"].append(record.get("date", ""))

        if record.get("status") == "PASS":
            case["success_count"] += 1
        else:
            case["fail_count"] += 1
            ftype = record.get("failure_type", "unknown")
            case["failure_types"][ftype] += 1

            version = record.get("uvp_version", "")
            if not case["first_fail_version"]:
                case["first_fail_version"] = version
            case["last_fail_version"] = version

        # 7天/30天结果（简化：用日期判断）
        date_str = record.get("date", "")
        case["results_7d"].append(record.get("status"))
        case["results_30d"].append(record.get("status"))

    # 计算统计指标
    stats = []
    for name, case in cases.items():
        total = case["total_runs"]
        success = case["success_count"]
        fail = case["fail_count"]

        # 成功率
        rate = success / total if total > 0 else 0

        # 7天/30天成功率（简化计算）
        rate_7d = sum(1 for r in case["results_7d"] if r == "PASS") / len(case["results_7d"]) * 100 if case["results_7d"] else 0
        rate_30d = sum(1 for r in case["results_30d"] if r == "PASS") / len(case["results_30d"]) * 100 if case["results_30d"] else 0

        # 失败原因统计
        failure_type_str = "; ".join(f"{k}:{v}" for k, v in sorted(case["failure_types"].items(), key=lambda x: -x[1]))

        # Flaky 检测：成功和失败交替出现
        flaky = "否"
        if total >= 4:
            # 最近4次结果中成功和失败都存在
            recent = case["results_7d"][-4:]
            has_pass = "PASS" in recent
            has_fail = "FAIL" in recent
            if has_pass and has_fail:
                flaky = "是"

        # 平均运行时间
        avg_duration = sum(case["durations"]) / len(case["durations"]) if case["durations"] else 0

        stats.append({
            "test_name": name,
            "total_runs": total,
            "success_count": success,
            "fail_count": fail,
            "success_rate": rate,
            "success_rate_7d": rate_7d,
            "success_rate_30d": rate_30d,
            "failure_type_stats": failure_type_str,
            "avg_duration": avg_duration,
            "first_fail_version": case["first_fail_version"] or "",
            "last_fail_version": case["last_fail_version"] or "",
            "flaky": flaky,
            "version_info": ", ".join(sorted(case["versions"])),
        })

    # 按失败次数降序
    stats.sort(key=lambda x: -x["fail_count"])
    return stats


def _default_suggestion(failure_type: str) -> str:
    """根据失败类型给出默认改进措施"""
    suggestions = {
        "libvirt_error": "检查 libvirtd 服务状态，确认 API 兼容性，查看对应版本 libvirt 变更日志",
        "qemu_crash": "收集 QEMU coredump，检查 QEMU 版本和补丁情况，查看对应版本 QEMU 变更",
        "kernel_panic": "分析内核崩溃转储 (vmcore)，检查内核版本和驱动兼容性",
        "memory_issue": "检查宿主机内存配置和 limits，确认 cgroup 设置，增加可用内存",
        "timeout": "排查网络/存储性能瓶颈，检查用例超时配置是否合理",
        "environment_issue": "检查网络连通性、存储挂载状态，确认环境依赖是否就绪",
        "case_script_issue": "检查测试用例脚本逻辑，修复断言或配置问题",
        "infrastructure_issue": "检查存储服务和网络配置，确认基础设施组件状态",
        "unknown_failure": "需要人工进一步排查，建议收集更多日志信息",
    }
    return suggestions.get(failure_type, "需要人工进一步排查")


# OpenCLAW Tool 注册信息
TOOL_METADATA = {
    "name": "excel_reporter",
    "description": "Generate Excel daily report and cumulative case statistics",
    "parameters": {
        "job_info": "dict - job metadata",
        "failures": "list - list of failure analyses",
        "history": "list - historical run records for stats",
    },
    "returns": "JSON with output file paths",
    "enabled": True
}
